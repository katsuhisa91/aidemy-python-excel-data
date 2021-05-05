import openpyxl

wb = openpyxl.load_workbook('売上報告書.xlsx', data_only=True)
sheet = wb['売上データ']

branch_list = list(sheet.columns)[2]
sales_per_customer = {}

# A支店の顧客別売上を集計する
for i, cell in enumerate(branch_list):
    if cell.value == "A支店":
        if sheet.cell(row=i + 1, column=2).value in sales_per_customer.keys():
            sales_per_customer[sheet.cell(row=i + 1, column=2).value] += \
                sheet.cell(row=i + 1, column=8).value
        else:
            sales_per_customer[sheet.cell(row=i + 1, column=2).value] = \
                sheet.cell(row=i + 1, column=8).value

# 売り上げの降順で並び替え
for k, v in sorted(sales_per_customer.items(), key=lambda x: -x[1]):
    print(k, v)
